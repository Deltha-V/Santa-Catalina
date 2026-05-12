from pathlib import Path
from tempfile import NamedTemporaryFile
import os
import uuid

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.core.config import settings
from app.core.security import require_admin
from app.db.session import get_db
from app.models.lote_imagen import LoteImagen
from app.models.lote import EstadoLote, Lote
from app.schemas.lote_imagen import LoteImagenRead
from app.schemas.lote import LoteCreate, LoteRead, LoteUpdate

router = APIRouter(prefix="/lotes", tags=["lotes"])


def _to_estado(raw: str):
    value = raw.strip().lower()
    if "disponible" == value:
        return EstadoLote.disponible
    if "vendido" in value:
        return EstadoLote.vendido
    if "reservado" in value:
        return EstadoLote.reservado
    if "acuerdos privados" in value or "fideicomiso" in value:
        return EstadoLote.acuerdo_privado
    if "no disponible" in value:
        return EstadoLote.no_disponible
    return EstadoLote.no_disponible


def _to_bool_comercializable(raw: str):
    value = raw.strip().lower()
    if "no comercializable" in value:
        return False
    return True


def _parse_medidas(raw: str):
    cleaned = raw.lower().replace(" ", "")
    if "x" not in cleaned:
        raise ValueError("Formato de medidas invalido")
    frente_raw, fondo_raw = cleaned.split("x", 1)
    frente = float(frente_raw.replace(",", "."))
    fondo = float(fondo_raw.replace(",", "."))
    return frente, fondo, frente * fondo


def _save_image(file: UploadFile, folder: str, filename_prefix: str) -> str:
    ext = Path(file.filename or "").suffix.lower()
    if ext not in {".png", ".jpg", ".jpeg", ".webp"}:
        raise HTTPException(status_code=400, detail="Formato de imagen no soportado")

    target_dir = Path(settings.uploads_dir) / folder
    target_dir.mkdir(parents=True, exist_ok=True)
    safe_name = f"{filename_prefix}-{uuid.uuid4().hex[:10]}{ext}"
    target = target_dir / safe_name
    with target.open("wb") as out:
        out.write(file.file.read())
    return f"/uploads/{folder}/{safe_name}"


@router.get("", response_model=list[LoteRead])
def list_lotes(db: Session = Depends(get_db)):
    return db.scalars(select(Lote).order_by(Lote.numero_lote.asc())).all()


@router.post("", response_model=LoteRead, status_code=status.HTTP_201_CREATED)
def create_lote(
    payload: LoteCreate,
    db: Session = Depends(get_db),
    _: None = Depends(require_admin),
):
    exists = db.get(Lote, payload.numero_lote)
    if exists:
        raise HTTPException(status_code=400, detail="El numero de lote ya existe")

    lote = Lote(**payload.model_dump())
    db.add(lote)
    db.commit()
    db.refresh(lote)
    return lote


@router.put("/{numero_lote}", response_model=LoteRead)
def update_lote(
    numero_lote: str,
    payload: LoteUpdate,
    db: Session = Depends(get_db),
    _: None = Depends(require_admin),
):
    lote = db.get(Lote, numero_lote)
    if not lote:
        raise HTTPException(status_code=404, detail="Lote no encontrado")

    for key, value in payload.model_dump().items():
        setattr(lote, key, value)

    db.commit()
    db.refresh(lote)
    return lote


@router.delete("/{numero_lote}", status_code=status.HTTP_204_NO_CONTENT)
def delete_lote(
    numero_lote: str,
    db: Session = Depends(get_db),
    _: None = Depends(require_admin),
):
    lote = db.get(Lote, numero_lote)
    if not lote:
        raise HTTPException(status_code=404, detail="Lote no encontrado")

    db.delete(lote)
    db.commit()
    return None


@router.post("/{numero_lote}/plano", response_model=LoteRead)
def upload_plano(
    numero_lote: str,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    _: None = Depends(require_admin),
):
    lote = db.get(Lote, numero_lote)
    if not lote:
        raise HTTPException(status_code=404, detail="Lote no encontrado")

    lote.plano_imagen_url = _save_image(file, "planos", numero_lote)
    db.commit()
    db.refresh(lote)
    return lote


@router.get("/{numero_lote}/imagenes", response_model=list[LoteImagenRead])
def list_lote_imagenes(numero_lote: str, db: Session = Depends(get_db)):
    lote = db.get(Lote, numero_lote)
    if not lote:
        raise HTTPException(status_code=404, detail="Lote no encontrado")
    query = (
        select(LoteImagen)
        .where(LoteImagen.lote_numero == numero_lote)
        .order_by(LoteImagen.orden.asc(), LoteImagen.id.asc())
    )
    return db.scalars(query).all()


@router.get("/imagenes/banco", response_model=list[LoteImagenRead])
def list_banco_imagenes(
    lote_numero: str | None = Query(default=None),
    db: Session = Depends(get_db),
):
    query = select(LoteImagen)
    if lote_numero:
        query = query.where(LoteImagen.lote_numero == lote_numero)
    query = query.order_by(LoteImagen.orden.asc(), LoteImagen.id.asc())
    return db.scalars(query).all()


@router.post("/{numero_lote}/imagenes", response_model=list[LoteImagenRead], status_code=status.HTTP_201_CREATED)
def upload_lote_imagenes(
    numero_lote: str,
    files: list[UploadFile] = File(...),
    db: Session = Depends(get_db),
    _: None = Depends(require_admin),
):
    lote = db.get(Lote, numero_lote)
    if not lote:
        raise HTTPException(status_code=404, detail="Lote no encontrado")
    if not files:
        raise HTTPException(status_code=400, detail="Debes subir al menos una imagen")

    last_order = db.scalar(
        select(LoteImagen.orden).where(LoteImagen.lote_numero == numero_lote).order_by(LoteImagen.orden.desc()).limit(1)
    )
    next_order = (last_order or 0) + 1

    created_ids: list[int] = []
    for item in files:
        image_url = _save_image(item, "lotes", numero_lote)
        image = LoteImagen(lote_numero=numero_lote, image_url=image_url, orden=next_order)
        next_order += 1
        db.add(image)
        db.flush()
        created_ids.append(image.id)

    db.commit()
    query = select(LoteImagen).where(LoteImagen.id.in_(created_ids)).order_by(LoteImagen.orden.asc(), LoteImagen.id.asc())
    return db.scalars(query).all()


@router.delete("/{numero_lote}/imagenes/{image_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_lote_imagen(
    numero_lote: str,
    image_id: int,
    db: Session = Depends(get_db),
    _: None = Depends(require_admin),
):
    image = db.get(LoteImagen, image_id)
    if not image or image.lote_numero != numero_lote:
        raise HTTPException(status_code=404, detail="Imagen no encontrada")

    if image.image_url.startswith("/uploads/"):
        rel_path = image.image_url.replace("/uploads/", "", 1)
        abs_path = Path(settings.uploads_dir) / rel_path
        if abs_path.exists():
            abs_path.unlink()

    db.delete(image)
    db.commit()
    return None


@router.post("/import-excel")
def import_lotes_from_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    _: None = Depends(require_admin),
):
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Sube un archivo .xlsx")

    inserted = 0
    updated = 0
    skipped = 0
    errors: list[str] = []

    with NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        tmp.write(file.file.read())
        tmp_path = tmp.name

    wb = load_workbook(tmp_path, data_only=True)
    try:
        for ws in wb.worksheets:
            if not ws.title.upper().startswith("MZ"):
                continue

            manzana = ws.title.replace("MZ", "").strip()
            for row in ws.iter_rows(min_row=2, max_col=13, values_only=True):
                lote_name = row[10] if len(row) > 10 else None  # col K
                medidas = row[11] if len(row) > 11 else None  # col L
                disponibilidad = row[12] if len(row) > 12 else None  # col M

                if not lote_name or not medidas or not disponibilidad:
                    continue
                if not str(lote_name).strip().upper().startswith("LOTE"):
                    skipped += 1
                    continue

                try:
                    lote_num = str(lote_name).strip().upper().replace("LOTE", "").strip()
                    frente_m, fondo_m, area_m2 = _parse_medidas(str(medidas))
                    estado = _to_estado(str(disponibilidad))
                    comercializable = _to_bool_comercializable(str(disponibilidad))
                    numero_lote = f"MZ{manzana}-L{lote_num}"
                except Exception as exc:
                    errors.append(f"{ws.title} {lote_name}: {str(exc)}")
                    skipped += 1
                    continue

                current = db.get(Lote, numero_lote)
                if current:
                    current.manzana = manzana
                    current.estado = estado
                    current.comercializable = comercializable
                    current.frente_m = frente_m
                    current.fondo_m = fondo_m
                    current.area_m2 = area_m2
                    updated += 1
                else:
                    db.add(
                        Lote(
                            numero_lote=numero_lote,
                            manzana=manzana,
                            estado=estado,
                            comercializable=comercializable,
                            frente_m=frente_m,
                            fondo_m=fondo_m,
                            area_m2=area_m2,
                        )
                    )
                    inserted += 1

        db.commit()
    finally:
        wb.close()
        if os.path.exists(tmp_path):
            os.unlink(tmp_path)

    return {
        "inserted": inserted,
        "updated": updated,
        "skipped": skipped,
        "errors_preview": errors[:15],
    }
